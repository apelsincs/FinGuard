"""
Сервис экспорта отчетов
Создает отчеты в различных форматах (Excel, PDF, CSV)
"""

import os
import csv
import io
from datetime import datetime, timedelta
from typing import List, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.database.models import Transaction, Category, Budget, TransactionType
from app.utils.logger import get_logger

logger = get_logger(__name__)


class ReportService:
    """Сервис для создания отчетов в различных форматах"""
    
    def __init__(self, db: Session):
        self.db = db
        self.reports_dir = "reports"
        
        # Создаем директорию для отчетов
        if not os.path.exists(self.reports_dir):
            os.makedirs(self.reports_dir)
    
    def export_to_excel(self, user_id: int, days: int = 30) -> str:
        """
        Экспортирует данные в Excel файл
        
        Args:
            user_id: ID пользователя
            days: Количество дней для анализа
            
        Returns:
            str: Путь к файлу отчета
        """
        try:
            # Получаем данные
            end_date = datetime.now()
            start_date = end_date - timedelta(days=days)
            
            transactions = self.db.query(Transaction).filter(
                Transaction.user_id == user_id,
                Transaction.transaction_date >= start_date,
                Transaction.transaction_date <= end_date
            ).order_by(Transaction.transaction_date.desc()).all()
            
            if not transactions:
                return None
            
            # Создаем DataFrame
            data = []
            for t in transactions:
                data.append({
                    'Дата': t.transaction_date.strftime('%d.%m.%Y %H:%M'),
                    'Тип': 'Доход' if t.type == TransactionType.INCOME else 'Расход',
                    'Сумма': t.amount,
                    'Валюта': t.currency,
                    'Описание': t.description or '',
                    'Категория': t.category.name if t.category else 'Без категории',
                    'Статус': t.status.value,
                    'Подозрительная': 'Да' if t.is_suspicious else 'Нет'
                })
            
            df = pd.DataFrame(data)
            
            # Создаем Excel файл
            wb = Workbook()
            ws = wb.active
            ws.title = "Транзакции"
            
            # Добавляем заголовок
            ws['A1'] = f"Финансовый отчет за {days} дней"
            ws['A1'].font = Font(size=16, bold=True)
            ws.merge_cells('A1:H1')
            
            # Добавляем дату создания
            ws['A2'] = f"Создан: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
            ws['A2'].font = Font(size=10, italic=True)
            ws.merge_cells('A2:H2')
            
            # Добавляем данные
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Стилизуем заголовки
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for cell in ws[4]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Добавляем итоги
            total_income = df[df['Тип'] == 'Доход']['Сумма'].sum()
            total_expense = df[df['Тип'] == 'Расход']['Сумма'].sum()
            balance = total_income - total_expense
            
            summary_row = len(df) + 6
            ws[f'A{summary_row}'] = "ИТОГО:"
            ws[f'A{summary_row}'].font = Font(bold=True)
            
            ws[f'C{summary_row}'] = f"Доходы: {total_income:,.0f} ₽"
            ws[f'C{summary_row}'].font = Font(bold=True, color="008000")
            
            ws[f'D{summary_row}'] = f"Расходы: {total_expense:,.0f} ₽"
            ws[f'D{summary_row}'].font = Font(bold=True, color="FF0000")
            
            ws[f'E{summary_row}'] = f"Баланс: {balance:+,.0f} ₽"
            ws[f'E{summary_row}'].font = Font(bold=True)
            
            # Настраиваем ширину столбцов
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Сохраняем файл
            filename = f"report_user_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.reports_dir, filename)
            wb.save(filepath)
            
            return filepath
            
        except Exception as e:
            logger.error(f"Ошибка при создании Excel отчета: {e}")
            return None
    
    def export_to_csv(self, user_id: int, days: int = 30) -> str:
        """
        Экспортирует данные в CSV файл
        
        Args:
            user_id: ID пользователя
            days: Количество дней для анализа
            
        Returns:
            str: Путь к файлу отчета
        """
        try:
            # Получаем данные
            end_date = datetime.now()
            start_date = end_date - timedelta(days=days)
            
            transactions = self.db.query(Transaction).filter(
                Transaction.user_id == user_id,
                Transaction.transaction_date >= start_date,
                Transaction.transaction_date <= end_date
            ).order_by(Transaction.transaction_date.desc()).all()
            
            if not transactions:
                return None
            
            # Создаем CSV файл
            filename = f"report_user_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            filepath = os.path.join(self.reports_dir, filename)
            
            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['Дата', 'Тип', 'Сумма', 'Валюта', 'Описание', 'Категория', 'Статус', 'Подозрительная']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                
                # Записываем заголовок
                writer.writeheader()
                
                # Записываем данные
                for t in transactions:
                    writer.writerow({
                        'Дата': t.transaction_date.strftime('%d.%m.%Y %H:%M'),
                        'Тип': 'Доход' if t.type == TransactionType.INCOME else 'Расход',
                        'Сумма': t.amount,
                        'Валюта': t.currency,
                        'Описание': t.description or '',
                        'Категория': t.category.name if t.category else 'Без категории',
                        'Статус': t.status.value,
                        'Подозрительная': 'Да' if t.is_suspicious else 'Нет'
                    })
            
            return filepath
            
        except Exception as e:
            logger.error(f"Ошибка при создании CSV отчета: {e}")
            return None
    
    def export_to_pdf(self, user_id: int, days: int = 30) -> str:
        """
        Экспортирует данные в PDF файл
        
        Args:
            user_id: ID пользователя
            days: Количество дней для анализа
            
        Returns:
            str: Путь к файлу отчета
        """
        try:
            # Получаем данные
            end_date = datetime.now()
            start_date = end_date - timedelta(days=days)
            
            transactions = self.db.query(Transaction).filter(
                Transaction.user_id == user_id,
                Transaction.transaction_date >= start_date,
                Transaction.transaction_date <= end_date
            ).order_by(Transaction.transaction_date.desc()).all()
            
            if not transactions:
                return None
            
            # Создаем PDF файл
            filename = f"report_user_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            filepath = os.path.join(self.reports_dir, filename)
            
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Заголовок
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # Центр
            )
            title = Paragraph(f"Финансовый отчет за {days} дней", title_style)
            story.append(title)
            
            # Дата создания
            date_style = ParagraphStyle(
                'CustomDate',
                parent=styles['Normal'],
                fontSize=10,
                spaceAfter=20,
                alignment=1
            )
            date_text = Paragraph(f"Создан: {datetime.now().strftime('%d.%m.%Y %H:%M')}", date_style)
            story.append(date_text)
            story.append(Spacer(1, 20))
            
            # Подготавливаем данные для таблицы
            table_data = [['Дата', 'Тип', 'Сумма', 'Описание', 'Категория']]
            
            for t in transactions:
                table_data.append([
                    t.transaction_date.strftime('%d.%m.%Y'),
                    'Доход' if t.type == TransactionType.INCOME else 'Расход',
                    f"{t.amount:,.0f} ₽",
                    t.description or '',
                    t.category.name if t.category else 'Без категории'
                ])
            
            # Создаем таблицу
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ALIGN', (2, 1), (2, -1), 'RIGHT'),  # Выравнивание сумм по правому краю
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
            
            # Добавляем итоги
            total_income = sum(t.amount for t in transactions if t.type == TransactionType.INCOME)
            total_expense = sum(t.amount for t in transactions if t.type == TransactionType.EXPENSE)
            balance = total_income - total_expense
            
            summary_style = ParagraphStyle(
                'Summary',
                parent=styles['Normal'],
                fontSize=12,
                spaceAfter=10
            )
            
            story.append(Paragraph(f"<b>ИТОГО:</b>", summary_style))
            story.append(Paragraph(f"Доходы: {total_income:,.0f} ₽", summary_style))
            story.append(Paragraph(f"Расходы: {total_expense:,.0f} ₽", summary_style))
            story.append(Paragraph(f"Баланс: {balance:+,.0f} ₽", summary_style))
            
            # Собираем PDF
            doc.build(story)
            
            return filepath
            
        except Exception as e:
            logger.error(f"Ошибка при создании PDF отчета: {e}")
            return None
    
    def create_monthly_report(self, user_id: int, year: int = None, month: int = None) -> str:
        """
        Создает месячный отчет
        
        Args:
            user_id: ID пользователя
            year: Год (по умолчанию текущий)
            month: Месяц (по умолчанию текущий)
            
        Returns:
            str: Путь к файлу отчета
        """
        try:
            if year is None:
                year = datetime.now().year
            if month is None:
                month = datetime.now().month
            
            # Определяем период
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = datetime(year, month + 1, 1) - timedelta(days=1)
            
            # Получаем данные
            transactions = self.db.query(Transaction).filter(
                Transaction.user_id == user_id,
                Transaction.transaction_date >= start_date,
                Transaction.transaction_date <= end_date
            ).order_by(Transaction.transaction_date.desc()).all()
            
            if not transactions:
                return None
            
            # Создаем Excel файл с дополнительной аналитикой
            wb = Workbook()
            
            # Лист с транзакциями
            ws_transactions = wb.active
            ws_transactions.title = "Транзакции"
            
            # Добавляем заголовок
            month_name = start_date.strftime('%B %Y')
            ws_transactions['A1'] = f"Месячный отчет: {month_name}"
            ws_transactions['A1'].font = Font(size=16, bold=True)
            ws_transactions.merge_cells('A1:H1')
            
            # Добавляем данные транзакций
            data = []
            for t in transactions:
                data.append({
                    'Дата': t.transaction_date.strftime('%d.%m.%Y'),
                    'Тип': 'Доход' if t.type == TransactionType.INCOME else 'Расход',
                    'Сумма': t.amount,
                    'Описание': t.description or '',
                    'Категория': t.category.name if t.category else 'Без категории',
                    'Статус': t.status.value
                })
            
            df = pd.DataFrame(data)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws_transactions.append(r)
            
            # Лист с аналитикой
            ws_analytics = wb.create_sheet("Аналитика")
            
            # Статистика по категориям
            category_stats = df.groupby(['Категория', 'Тип'])['Сумма'].sum().reset_index()
            
            ws_analytics['A1'] = "Статистика по категориям"
            ws_analytics['A1'].font = Font(size=14, bold=True)
            
            # Добавляем данные категорий
            for i, (_, row) in enumerate(category_stats.iterrows(), 3):
                ws_analytics[f'A{i}'] = row['Категория']
                ws_analytics[f'B{i}'] = row['Тип']
                ws_analytics[f'C{i}'] = row['Сумма']
            
            # Заголовки для статистики
            ws_analytics['A2'] = "Категория"
            ws_analytics['B2'] = "Тип"
            ws_analytics['C2'] = "Сумма"
            
            for cell in ws_analytics[2]:
                cell.font = Font(bold=True)
            
            # Сохраняем файл
            filename = f"monthly_report_user_{user_id}_{year}_{month:02d}.xlsx"
            filepath = os.path.join(self.reports_dir, filename)
            wb.save(filepath)
            
            return filepath
            
        except Exception as e:
            logger.error(f"Ошибка при создании месячного отчета: {e}")
            return None
    
    def cleanup_old_reports(self, max_age_hours: int = 24):
        """
        Удаляет старые отчеты
        
        Args:
            max_age_hours: Максимальный возраст файлов в часах
        """
        try:
            cutoff_time = datetime.now() - timedelta(hours=max_age_hours)
            
            for filename in os.listdir(self.reports_dir):
                if filename.endswith(('.xlsx', '.csv', '.pdf')):
                    filepath = os.path.join(self.reports_dir, filename)
                    file_time = datetime.fromtimestamp(os.path.getctime(filepath))
                    
                    if file_time < cutoff_time:
                        os.remove(filepath)
                        logger.info(f"Удален старый отчет: {filename}")
                        
        except Exception as e:
            logger.error(f"Ошибка при очистке старых отчетов: {e}")
